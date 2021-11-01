import pandas as pd
from threading import Lock
import openpyxl
import os

class SaveResult():
    def __init__(self, number_agent, path = "results.xlsx"):
        self.data_lock = Lock()
        self.path_result = path
        self.name_cur_sheet = ""
        self.number_agent = number_agent
        self.number_parameters = 6
        self.init_sheet()

    def init_sheet(self):
        number = 0
        if os.path.isfile(self.path_result):
            exist_wb = openpyxl.load_workbook(self.path_result)
            number = len(exist_wb.sheetnames)
        wb = openpyxl.Workbook()
        wb.create_sheet(title = str(number), index = number+1)
        sheet = wb[str(number)]

        for n in range(self.number_agent):
            cell_loss = sheet.cell(row=1, column = self.number_parameters * n + 2)
            cell_loss.value = "Loss Agent {}" .format(n)
            cell_reward = sheet.cell(row=1, column = self.number_parameters * n + 3)
            cell_reward.value = "Reward Agent {}" .format(n)
            cell_reward = sheet.cell(row=1, column = self.number_parameters * n + 4)
            cell_reward.value = "TP".format(n)
            cell_reward = sheet.cell(row=1, column = self.number_parameters * n + 5)
            cell_reward.value = "TN".format(n)
            cell_reward = sheet.cell(row=1, column = self.number_parameters * n + 6)
            cell_reward.value = "FP".format(n)
            cell_reward = sheet.cell(row=1, column = self.number_parameters * n + 7)
            cell_reward.value = "FN".format(n)

        wb.save(self.path_result)
        self.name_cur_sheet = str(number)

    def save_loss(self, idx, loss):
        with self.data_lock:
            wb = openpyxl.load_workbook(self.path_result)
            sheet = wb[self.name_cur_sheet]
            rows = sheet.max_row
            if idx == 0:
                rows += 1
            cell = sheet.cell(row = rows, column = self.number_parameters * idx+2)
            cell.value = loss.tolist()
            wb.save(self.path_result)


    def save_reward(self, idx, reward):
        with self.data_lock:
            wb = openpyxl.load_workbook(self.path_result)
            sheet = wb[self.name_cur_sheet]
            rows = sheet.max_row
            cell = sheet.cell(row = rows, column = self.number_parameters * idx+3)
            cell.value = reward
            wb.save(self.path_result)

    def save_metrics(self, idx, fp, fn, tp, tn):
        with self.data_lock:
            wb = openpyxl.load_workbook(self.path_result)
            sheet = wb[self.name_cur_sheet]
            rows = sheet.max_row

            cell = sheet.cell(row = rows, column = self.number_parameters * idx+4)
            cell.value = tp

            cell = sheet.cell(row=rows, column= self.number_parameters * idx + 5)
            cell.value = tn

            cell = sheet.cell(row=rows, column= self.number_parameters * idx + 6)
            cell.value = fp

            cell = sheet.cell(row=rows, column=self.number_parameters * idx + 7)
            cell.value = fn
            wb.save(self.path_result)

import os
from random import random, randint
from mlflow import log_metric, log_param, log_artifacts

if __name__ == "__main__":
    # Log a parameter (key-value pair)
    log_param("param1", randint(0, 100))

    # Log a metric; metrics can be updated throughout the run
    log_metric("foo", random())
    log_metric("foo", random() + 1)
    log_metric("foo", random() + 2)

    # Log an artifact (output file)
    if not os.path.exists("outputs"):
        os.makedirs("outputs")
    with open("outputs/test.txt", "w") as f:
        f.write("hello world!")
    log_artifacts("outputs")
